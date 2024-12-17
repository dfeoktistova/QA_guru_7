import requests
import os
import zipfile
import pytest
import shutil
from pypdf import PdfReader
import csv
from openpyxl import load_workbook


CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
TMP_DIR = os.path.join(CURRENT_DIR, "download")

files_url = {
        'pdf':  'https://example-files.online-convert.com/document/pdf/example.pdf',
        'xlsx': 'https://example-files.online-convert.com/spreadsheet/xlsx/example.xlsx',
        'csv': 'https://example-files.online-convert.com/spreadsheet/csv/example.csv'
    }


@pytest.fixture(autouse=True, scope='function')
def setup():
    if not os.path.exists("download"):
        os.mkdir("download")

    yield

    shutil.rmtree(f'{CURRENT_DIR}/download')


@pytest.fixture(autouse=True, scope='function')
def download_files(setup):
    for file_url in files_url:
        content = requests.get(url=files_url[file_url]).content
        file_name = f'{file_url}.{file_url}'
        with open(f'{TMP_DIR}/{file_name}', 'wb') as file:
            file.write(content)


@pytest.fixture(autouse=True, scope='function')
def create_zip(setup, download_files):
    files = os.listdir(TMP_DIR)

    with zipfile.ZipFile(f"{TMP_DIR}/zip.zip", mode="w") as archive:
        for file in files:
            new_file = os.path.join(TMP_DIR, file)
            archive.write(new_file, os.path.basename(new_file))
            archive.printdir()


def get_zip_data(file_name):
    with zipfile.ZipFile(f"{TMP_DIR}/zip.zip", mode="r") as archive:
        with archive.open(f'{file_name}.{file_name}') as file:
            if file_name == 'pdf':
                reader = PdfReader(file)
                text = ''
                for page in reader.pages:
                    text += f'{page.extract_text()} \n'
                return text

            elif file_name == 'xlsx':
                workbook = load_workbook(file)
                sheet = workbook.active
                text = ''
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        text += f'{cell.value} \n'
                return text

            elif file_name == 'csv':
                content = file.read().decode()
                csv_reader = list(csv.reader(content.splitlines()))
                text = ''
                for line in csv_reader:
                    text += ','.join(line) + '\n'
                return text